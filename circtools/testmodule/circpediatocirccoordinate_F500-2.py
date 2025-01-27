import os
import re
import requests
import json
import time
import openpyxl
import statistics

from openpyxl.styles import PatternFill
from datetime import datetime

#from circtools.testmodule.circpediatocirccoordinate import chromosome

# initializing parameters


filepath = "/home/bjamshidikia/circtoolsdata"
url = "https://circhemy.jakobilab.org/api/query"
headers = {
    "accept": "application/json",
    "Content-Type": "application/json"
}

if os.path.isfile(filepath + "/CircCoordinates" ) == True:
    CircCoordinate = open(filepath + "/CircCoordinates","r")
else:
        print("File CircCoordinate not found ")
        exit()

listtimes = []
listrecords = []
start_run = datetime.now().strftime("%H:%M:%S")

for z in range (0,10):

    line = CircCoordinate.readline()    # sending the file pointer to the second line for the first read
    #print(line)



    # writing header line
    #w = open(filepath + "/circpediacircCoordinte.bed", "w") #open(filepath + "/CountRNA-" + outfilelist[j]+".bed", "w") # openning the appropriate out file
    spline = line.split()
#    wlist = spline[0:8]
#    wlist.append("Circpedia2")
#    wline = "\t".join(wlist)
    #w.write(wline + "\n")

    line = CircCoordinate.readline()
    spline = line.split()       # making a list from the line
    chromosome = "chr" + spline[0]  # "chr22"
    chrstart = spline[1]  # "36341370"
    chrstop = spline[2]  # "36349255"

    data = '{"input": [{"query": "' + chromosome + '", "field": "Chr", "operator1": "AND", "operator2": "is"},' \
           '{"query": "' + chrstart + '", "field": "Start", "operator1": "AND", "operator2": "is"},' \
           '{"query": "' + chrstop + '", "field": "Stop", "operator1": "AND", "operator2": "is"}'



    lsreadlines = []
    lsreadlines.append([spline[0:8]])

    start_time = time.time()



    for i in range(0,500):

        line = CircCoordinate.readline() # sending the file pointer to the second line for every other reads
    #    print ( "i = " + str(i))
        if not line :
            break
        spline = line.split()       # making a list from the line
        #wlist = spline[0:8]         # adding first four columns chr,start,end,Gene name
        chromosome = "chr" + spline[0]  #"chr22"
        chrstart = spline[1]#"36341370"
        chrstop = spline[2] #"36349255"
    #    print("chr : "+ chromosome + " Start : " + chrstart + " Stop : " + chrstop  )

        data1 =  ',{"query": "' + chromosome + '", "field": "Chr", "operator1": "OR", "operator2": "is"},{"query": "' + chrstart + '", "field": "Start", "operator1": "AND", "operator2": "is"},{"query": "' + chrstop + '", "field": "Stop", "operator1": "AND", "operator2": "is"}'
        data = data + data1
    #    print(data)
        lsreadlines.append([spline[0:8]])
        i +=  1

    data = data + '],'

    data = data + ' "output": ["Circpedia2","Chr","Start","Stop"]}'

    response = requests.post(url, data=data, headers=headers)
         # print(response.json())
    jsondata = response.json()
    totaltime = time.time() - start_time
    print("time for the :"+ str(z) + " read is : " + str(totaltime))
    #print(jsondata)
    #print("total number of records : " + str(len(jsondata["rowData"])))
    listtimes.append(totaltime)
    listrecords.append(len(jsondata["rowData"]))
    CircCoordinate.seek(0,0)
    lsreadlines = []
    jsondata ={}

z += 1
#jsondata["rowData"]


"""
for i in range(0,len(lsreadlines)):
#    print(" i = " + str(i))
    chromosome = "chr" + str(lsreadlines[i][0][0])
    chrstart = lsreadlines[i][0][1]
    chrstop= lsreadlines[i][0][2]
    #print(chromosome + " , " + chrstart + " , "+ chrstop)
#    print("--------------")
    for j in range(0,len(jsondata["rowData"])):
        jj = jsondata["rowData"][j]#["Circpedia2"]
        #print( str(jsondata["rowData"][j]["Chr"]) + " , " + str(jsondata["rowData"][j]["Start"]) + " , " + str(jsondata["rowData"][j]["Stop"]))
        if str(jsondata["rowData"][j]["Chr"]).strip() == chromosome.strip() and str(jsondata["rowData"][j]["Start"]).strip() == chrstart.strip() and str(jsondata["rowData"][j]["Stop"]).strip() == chrstop.strip() :
            #print( " circpedia2: "+ str(jsondata["rowData"][j]["Circpedia2"]))
            strcircpedia = str(jsondata["rowData"][j]["Circpedia2"])
        #else:
        #    print("None")
        j +=1
    lsreadlines[i][0].append(strcircpedia)
    wline = "\t".join(lsreadlines[i][0])
    print(wline)
    w.write( wline + "\n" )     # write the tab delimited string in the appropriate file
    strcircpedia = ""

    i += i
"""
wb = openpyxl.load_workbook("/home/bjamshidikia/speedtest_2.xlsx")
sheet = wb.active

sheet["E1"] = start_run
sheet["D1"] = "start time "


sheet["E2"] = "Times for 500 "
sheet["D2"] = "Number of records"

cellbg1 = sheet["E2"]
cellbg2 = sheet["D2"]
cellbg1.fill = PatternFill(start_color="ffcc00",end_color="ffcc00",fill_type="solid")
cellbg2.fill = PatternFill(start_color="ffcc00",end_color="ffcc00",fill_type="solid")

for i in range (0,len(listtimes)):
# Write data to the sheet
    sheet["E"+str(i+3)] = listtimes[i]
    sheet["D"+str(i+3)] = listrecords[i]

sheet["E"+str(str(len(listtimes)+3))] = statistics.mean(listtimes)
cellbg3 = sheet["E"+str(str(len(listtimes)+3))]
cellbg3.fill = PatternFill(start_color="ff0000",end_color="ff0000",fill_type="solid")

sheet["D"+str(str(len(listtimes)+3))] = "Average"
cellbg4 = sheet["D"+str(str(len(listtimes)+3))]
cellbg4.fill = PatternFill(start_color="ff0000",end_color="ff0000",fill_type="solid")



print(statistics.mean(listtimes))
# Save the workbook
wb.save("/home/bjamshidikia/speedtest_2.xlsx")
wb.close()

#w.close()               #closing the last output file
CircCoordinate.close()  # closing the input file CircCoordinate



