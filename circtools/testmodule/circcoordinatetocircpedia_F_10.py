import os
import re
import requests
import json
import time
import openpyxl
from openpyxl.styles import PatternFill
import statistics
#from circtools.testmodule.circpediatocirccoordinate import chromosome

# initializing parameters


filepath = "/home/bjamshidikia/circtoolsdata"
url = "https://circhemy.jakobilab.org/api/query"
headers = {
    "accept": "application/json",
    "Content-Type": "application/json"
}

if os.path.isfile(filepath + "/CircCoordinates" ) == True:
    CircCoordinate = open(filepath + "/CircCoordinates","r")
else:
        print("File CircCoordinate not found ")
        exit()

listtimes = []
listsubtime = []
listrecords = []


for z in range (0,1):
    start_time = time.time()
    line = CircCoordinate.readline()  # sending the file pointer to the second line for the first read
    spline = line.split()
    for t in range(0,33):
        time1 = time.time()
            #print(line)
            # writing header line
            #w = open(filepath + "/circpediacircCoordinte.bed", "w") #open(filepath + "/CountRNA-" + outfilelist[j]+".bed", "w") # openning the appropriate out file
        #    wlist = spline[0:8]
        #    wlist.append("Circpedia2")
        #    wline = "\t".join(wlist)
            #w.write(wline + "\n")

        line = CircCoordinate.readline()
        if not line:
            break

 #       print(line)
        spline = line.split()       # making a list from the line
        chromosome = "chr" + spline[0]  # "chr22"
        chrstart = spline[1]  # "36341370"
        chrstop = spline[2]  # "36349255"

        data = '{"input": [{"query": "' + chromosome + '", "field": "Chr", "operator1": "AND", "operator2": "is"},' \
               '{"query": "' + chrstart + '", "field": "Start", "operator1": "AND", "operator2": "is"},' \
               '{"query": "' + chrstop + '", "field": "Stop", "operator1": "AND", "operator2": "is"}'
        lsreadlines = []
        lsreadlines.append([spline[0:8]])
        for i in range(0,9):

            line = CircCoordinate.readline() # sending the file pointer to the second line for every other reads
#            print("------  " + line)
        #    print ( "i = " + str(i))
            if not line :
                break
            spline = line.split()       # making a list from the line
            #wlist = spline[0:8]         # adding first four columns chr,start,end,Gene name
            chromosome = "chr" + spline[0]  #"chr22"
            chrstart = spline[1]#"36341370"
            chrstop = spline[2] #"36349255"
        #    print("chr : "+ chromosome + " Start : " + chrstart + " Stop : " + chrstop  )

            data1 =  ',{"query": "' + chromosome + '", "field": "Chr", "operator1": "OR", "operator2": "is"},{"query": "' + chrstart + '", "field": "Start", "operator1": "AND", "operator2": "is"},{"query": "' + chrstop + '", "field": "Stop", "operator1": "AND", "operator2": "is"}'
            data = data + data1
        #    print(data)
            lsreadlines.append([spline[0:8]])
            i +=  1
        i = 0
        data = data + '],'

        data = data + ' "output": ["Circpedia2","Chr","Start","Stop"]}'

        response = requests.post(url, data=data, headers=headers)
        jsondata = response.json()
        listrecords.append(len(jsondata["rowData"]))
        if len(jsondata["rowData"]) == 15:
            print(data)
            print("--------")
            print(jsondata["rowData"])
            exit()
        subtime = time.time()-time1
        listsubtime.append(subtime)
#        print ("z =  " + str(z) + " t = " + str(t) + " Time = " + str(subtime) )
        jsondata = {}
        t +=1
             # print(response.json())

    totaltime = time.time() - start_time
    print("time for the :"+ str(z) + " read is : " + str(totaltime))
    #print(jsondata)
    #print("total number of records : " + str(len(jsondata["rowData"])))
    listtimes.append(totaltime)
    CircCoordinate.seek(0,0)
    lsreadlines = []
    jsondata ={}

z += 1
#jsondata["rowData"]


"""
for i in range(0,len(lsreadlines)):
#    print(" i = " + str(i))
    chromosome = "chr" + str(lsreadlines[i][0][0])
    chrstart = lsreadlines[i][0][1]
    chrstop= lsreadlines[i][0][2]
    #print(chromosome + " , " + chrstart + " , "+ chrstop)
#    print("--------------")
    for j in range(0,len(jsondata["rowData"])):
        jj = jsondata["rowData"][j]#["Circpedia2"]
        #print( str(jsondata["rowData"][j]["Chr"]) + " , " + str(jsondata["rowData"][j]["Start"]) + " , " + str(jsondata["rowData"][j]["Stop"]))
        if str(jsondata["rowData"][j]["Chr"]).strip() == chromosome.strip() and str(jsondata["rowData"][j]["Start"]).strip() == chrstart.strip() and str(jsondata["rowData"][j]["Stop"]).strip() == chrstop.strip() :
            #print( " circpedia2: "+ str(jsondata["rowData"][j]["Circpedia2"]))
            strcircpedia = str(jsondata["rowData"][j]["Circpedia2"])
        #else:
        #    print("None")
        j +=1
    lsreadlines[i][0].append(strcircpedia)
    wline = "\t".join(lsreadlines[i][0])
    print(wline)
    w.write( wline + "\n" )     # write the tab delimited string in the appropriate file
    strcircpedia = ""

    i += i
"""
wb = openpyxl.load_workbook("/home/bjamshidikia/speedtest.xlsx")
sheet = wb.active

sheet["G1"] = "Number of records"
sheet["H1"] = "Time / 10 records"
sheet["I1"] = "time / 326"

cellbg1 = sheet["G1"]
cellbg2 = sheet["H1"]
cellbg3 = sheet["I1"]
cellbg1.fill = PatternFill(start_color="ffcc00",end_color="ffcc00",fill_type="solid")
cellbg2.fill = PatternFill(start_color="ffcc00",end_color="ffcc00",fill_type="solid")
cellbg3.fill = PatternFill(start_color="ffcc00",end_color="ffcc00",fill_type="solid")



for i in range (0,len(listsubtime)):
# Write data to the sheet
    sheet["G"+str(i+2)] = listrecords[i]
    sheet["H" + str(i+2)] = listsubtime[i]
for j in range (0,len(listtimes)):
    sheet["I"+str(j+2)] = listtimes[j]

avg1 = statistics.mean(listsubtime)
avg2 = statistics.mean(listtimes)

#sheet["G" + str(i + 2)] = "average"
#sheet["G" + str(j + 2)] = "average"
cellbg4 = sheet["G"+ str(i+2)]
cellbg5 = sheet["G" + str(j+2)]

cellbg4.fill = PatternFill(start_color="ff0000",end_color="ff0000",fill_type="solid")
cellbg5.fill = PatternFill(start_color="ff0000",end_color="ff0000",fill_type="solid")

sheet["H"+ str(i+2)] = avg1
sheet["I"+ str(j+2)] = avg2
cellbg1 = sheet["G"+ str(i+2)]
cellbg2 = sheet["H" + str(i+2)]
cellbg3 = sheet["I" +str(j+2)]

cellbg1.fill = PatternFill(start_color="ff0000",end_color="ff0000",fill_type="solid")
cellbg2.fill = PatternFill(start_color="ff0000",end_color="ff0000",fill_type="solid")
cellbg3.fill = PatternFill(start_color="ff0000",end_color="ff0000",fill_type="solid")







# Save the workbook
wb.save("/home/bjamshidikia/speedtest.xlsx")


#w.close()               #closing the last output file
CircCoordinate.close()  # closing the input file CircCoordinate



